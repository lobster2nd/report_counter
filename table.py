import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment
import flet as ft

from fields import values

file_path = 'отчёт.xlsx'

CELL_VALUES = {
            'A1': 'Количество процедур по исследованиям',
            'A5': 'Наименование',
            'A6': 'Всего',
            'A7': 'ОГК',
            'A8': 'Костно-мышечной системы',
            'A9': 'Конечности',
            'A10': 'Таза и тазобедренных суставов',
            'A11': 'Шейные позвонки',
            'A12': 'Грудные позвонки',
            'A13': 'Поясничные позвонки',
            'A14': 'Рёбра и грудина',
            'A15': 'Черепа и челюстно-лицевой области',
            'A16': 'Зубы',
            'A17': 'Челюстей',
            'A18': 'Околоносовых пазух',
            'A19': 'Череп',
            'A20': 'Брюшная полость',
            'A22': 'Сохранено: ',
            'B5': 'Всего исследований',
            'C5': 'Всего снимков',
            'B6': '=B7 + B8 + B15',
            'C6': '=C7 + C8 + C15',
            'B8': '=SUM(B9:B14)',
            'C8': '=SUM(C9:C14)',
            'B15': '=SUM(B16:B20)',
            'C15': '=SUM(C16:C20)',
            'B22': f'{datetime.now().strftime("%d-%m-%y %H:%M")}'
        }


def clear_fields(e, page):
    """Очистить поля"""
    for v in values:
        v[1].value = None
        v[2].value = None
    page.update()


def create_table():
    """Создаёт шаблон таблицы отчёта"""
    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.column_dimensions['A'].width = 40
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20

        for cell, formula in CELL_VALUES.items():
            sheet[cell].value = formula

        for row in sheet.iter_rows(min_row=5, max_row=20, min_col=2,
                                   max_col=3):
            for cell in row:
                if cell.value is None:
                    cell.value = 0

        for col in ['B', 'C']:
            for cell in sheet[col]:
                cell.alignment = Alignment(horizontal='center',
                                           vertical='center')

        wb.save(file_path)


def update_table_values(action):
    """Обновляет или перезаписывает значения в таблице"""

    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    for val in values:
        section_name = val[0].value
        scan_cnt_str = val[1].value
        img_cnt_str = val[2].value

        try:
            scan_cnt = int(scan_cnt_str)
            img_cnt = int(img_cnt_str)
        except ValueError:
            continue

        for key, formula in CELL_VALUES.items():
            if section_name in formula and 'области' not in formula:
                cell_b = sheet['B' + key[1:]]
                cell_c = sheet['C' + key[1:]]

                if action == 'add':
                    cell_b.value += scan_cnt
                    cell_c.value += img_cnt
                elif action == 'rewrite':
                    cell_b.value = scan_cnt
                    cell_c.value = img_cnt

    wb.save(file_path)

    for v in values:
        v[1].value = None
        v[2].value = None


def add_to_table_values(e, page):
    """Прибавить переданные значения к таблице"""
    page.snack_bar = ft.SnackBar(ft.Text('Сохранено'))
    page.snack_bar.open = True

    update_table_values(action='add')
    page.update()


def rewrite_table_values(e, page):
    """Внести новые значения в таблицу, старые удаляются"""
    page.snack_bar = ft.SnackBar(ft.Text('Сохранено'))
    page.snack_bar.open = True

    update_table_values(action='rewrite')
    page.update()
